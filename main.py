from openpyxl import load_workbook

'''скрипт для поиска виниловых пластинок по каталогу testhifi'''

wb = load_workbook('/home/pixelastra/Документы/list_lp.xlsx') # открытие книги
data = wb['LP - виниловые пластинки']  # создание объекта с данными страницы LP...
desired_price = 900  # переменная с указанием цены, меньше которой будут показываться пластинки
counter = 1  # объявление счетчика
final_value = 100000  # объявление переменной, в которой будут храниться значение цены
for i in range(data.max_row): # приводим все значения с разделителями ., (str) к типу int
    position = 'K' + str(i + 3)  # каталог всегда начинается с 3 строки, проверка начнется с нее
    if isinstance(data[position].value, str): # если тип данных ячейки str - сделать преобразование
        if '.' in data[position].value[:3]: # случай, если точка - разделитель
            value_commas = data[position].value.partition(',')[0]  # удаление символов после запятой (до копеек)
            value_int = int(value_commas.replace('.', ''))  # удаление разделительной '.', преобразование в тип int
            final_value = value_int  # запись значения в переменную
        elif ',' in data[position].value[:3]:  # случай, если запятая - разделитель
            value_commas = data[position].value.partition('.')[0]  # удаление символов после точки
            value_int = int(value_commas.replace(',', ''))  # удаление разделительной ',', преобразование в тип int
            final_value = value_int  # запись значения в переменную
        else:  # ошибки при заполнении, например, цена перепутана с магазином
            print(
                "WARNING: In position {0} in text ({1}) something's went wrong".format(position, data[position].value))
    if isinstance(data[position].value, int):  # если типом переменной является int преобразования не нужны
        final_value = data[position].value
    if data[position].value is None: # пропустить, если цены нет
        continue
    if final_value <= desired_price:  # проверка условия "цена пластинки меньше желаемой"
        output = position + ''
        for c in range(1, 13):  # вывод столбцов с 1 по 12
            output += ' / ' + str(data.cell(row=i+3,column=c).value)
            output.replace('\n', '')
        print(output + '\n')  # Разделить вывод
        counter += 1
